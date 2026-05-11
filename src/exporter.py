from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import config


def export_to_excel(model, staff_data, shift_data, output_file=None):
   
    if output_file is None:
        output_file = config.OUTPUT_SCHEDULE_FILE
    
    if model.solution is None or model.solution.empty:
        raise ValueError("No solution to export. Model might not be solved.")
    
    solution_with_details = model.solution.merge(
        shift_data[['UNIQUE_KEY', 'MS_CA', 'Cơ sở', 'Ca thi', 'Ngày', 'Giờ', 'Thứ']],
        on='UNIQUE_KEY',
        how='left'
    )
    
    wb = Workbook()
    wb.remove(wb.active)

    _create_shift_schedule_sheet(wb, solution_with_details, shift_data)
    _create_staff_schedule_sheet(wb, solution_with_details, staff_data)
    _create_summary_sheet(wb, model, staff_data, shift_data)
    
    wb.save(output_file)
    print(f"OK Results exported to: {output_file}")
    
    return output_file


def _write_headers(ws, headers, fill_color):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")


def _create_shift_schedule_sheet(wb, solution_with_details, shift_data):
    """Create sheet showing which staff are assigned to each shift."""
    ws = wb.create_sheet("Ca Thi")

    headers = ['MS_CA', 'Ca thi', 'Ngày', 'Giờ', 'Cơ sở', 'Thứ', 'Số người', 'Danh sách cán bộ']
    _write_headers(ws, headers, "4472C4")

    row_idx = 2
    for _, shift_row in shift_data.iterrows():
        unique_key = shift_row['UNIQUE_KEY']
        assigned_staff = solution_with_details[solution_with_details['UNIQUE_KEY'] == unique_key]['MS_CB'].tolist()
        
        ws.cell(row=row_idx, column=1).value = shift_row.get('MS_CA', '')
        ws.cell(row=row_idx, column=2).value = shift_row.get('Ca thi', '')
        ws.cell(row=row_idx, column=3).value = shift_row.get('Ngày', '')
        ws.cell(row=row_idx, column=4).value = shift_row.get('Giờ', '')
        ws.cell(row=row_idx, column=5).value = shift_row.get('Cơ sở', '')
        ws.cell(row=row_idx, column=6).value = shift_row.get('Thứ', '')
        ws.cell(row=row_idx, column=7).value = len(assigned_staff)
        ws.cell(row=row_idx, column=8).value = ', '.join(assigned_staff)
        
        row_idx += 1
    
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.column_dimensions['H'].width = 30
    
    return ws


def _create_staff_schedule_sheet(wb, solution_with_details, staff_data):
    """Create sheet showing which shifts are assigned to each staff."""
    ws = wb.create_sheet("Cán Bộ")

    headers = ['MS_CB', 'Tuổi', 'Giới tính', 'Số ca trực', 'Danh sách ca', 'KC CS1 (km)', 'KC CS2 (km)']
    _write_headers(ws, headers, "70AD47")

    assignment_map = (
        solution_with_details.assign(
            shift_label=solution_with_details['MS_CA'] + ' (' + solution_with_details['Cơ sở'] + ')'
        )
        .groupby('MS_CB')['shift_label']
        .agg(', '.join)
        .to_dict()
    )

    row_idx = 2
    for _, staff_row in staff_data.iterrows():
        cb = staff_row['MS_CB']
        assigned_shifts = assignment_map.get(cb, '')

        ws.cell(row=row_idx, column=1).value = cb
        ws.cell(row=row_idx, column=2).value = staff_row.get('Tuổi', '')
        ws.cell(row=row_idx, column=3).value = staff_row.get('Giới tính', '')
        ws.cell(row=row_idx, column=4).value = len(assigned_shifts.split(', ')) if assigned_shifts else 0
        ws.cell(row=row_idx, column=5).value = assigned_shifts
        ws.cell(row=row_idx, column=6).value = staff_row.get('KC CS1 (km)', '')
        ws.cell(row=row_idx, column=7).value = staff_row.get('KC CS2 (km)', '')
        row_idx += 1

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.column_dimensions['E'].width = 30
    return ws


def _create_summary_sheet(wb, model, staff_data, shift_data):
    """Create sheet with summary statistics."""
    ws = wb.create_sheet("Thống Kê")
    
    stats = model.get_summary_stats()
    
    # Title
    ws.cell(row=1, column=1).value = "Thống Kê Lịch Trực"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    # Statistics
    row = 3
    summary_data = [
        ('Tổng số ca trực:', stats['total_shifts']),
        ('Tổng số cán bộ:', stats['total_staff']),
        ('Tổng số lần gán:', stats['total_assignments']),
        ('Trung bình người/ca:', f"{stats['average_staff_per_shift']:.2f}"),
        ('Ca nhiều nhất/người:', stats['max_shifts_per_staff']),
        ('Ca ít nhất/người:', stats['min_shifts_per_staff']),
        ('Trung bình ca/người:', f"{stats['avg_shifts_per_staff']:.2f}"),
    ]
    
    for label, value in summary_data:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    
    return ws


if __name__ == '__main__':
    from .loader import load_staff_data, load_shift_data
    from .model import ExamSchedulerModel
    
    staff = load_staff_data()
    shifts = load_shift_data()
    
    model = ExamSchedulerModel(staff, shifts)
    model.create_model()
    model.solve()
    model.extract_solution()
    
    export_to_excel(model, staff, shifts)